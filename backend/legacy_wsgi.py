# pyright: reportGeneralTypeIssues=false
# =========================================================
# DOC MANAGEMENT SYSTEM — MAIN FLASK APPLICATION (Mode 1)
# =========================================================
from sqlalchemy.orm import joinedload  # 👈 put this at the top of the file
from dotenv import load_dotenv
import os

load_dotenv()
import os
import sys
import pymysql
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect,
    url_for, flash, session, jsonify, Blueprint, send_file, abort
)
from sqlalchemy import exists
from sqlalchemy import func, inspect
from sqlalchemy.exc import OperationalError
from sqlalchemy.dialects import sqlite as sqlite_dialect
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.utils import secure_filename
from sqlalchemy import case
from sqlalchemy import cast, Integer


from extensions import db
from openpyxl import Workbook

from urllib.error import URLError, HTTPError
from models import Entity,Series, Subseries, Doctype

from urllib.parse import urlparse
from datetime import datetime
from datetime import time as dt_time

from sqlalchemy import UniqueConstraint
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from tempfile import NamedTemporaryFile
from urllib.request import urlopen, Request

#from flask import send_file, request, redirect, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from copy import copy
from collections import OrderedDict
from flask import render_template
from collections import OrderedDict

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from flask import send_file
from io import BytesIO
from datetime import datetime
import os

# =========================================================
# BACKEND – PROFESSIONAL XLS EXPORT
# =========================================================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from flask import send_file
from io import BytesIO
from datetime import datetime
from flask import make_response

import os

# =========================================================
# BASE PATHS AND DIRECTORIES
# =========================================================
def get_base_path():
    """Return the base path for templates/static, compatible with PyInstaller."""
    if getattr(sys, "frozen", False):
        return sys._MEIPASS  # type: ignore[attr-defined]
    return os.path.abspath(".")

base_path = get_base_path()

if sys.platform == "win32":
    user_data_dir = os.path.join(os.environ.get("APPDATA", "."), "DocMng")
else:
    user_data_dir = os.path.join(os.path.expanduser("~"), ".DocMng")

os.makedirs(user_data_dir, exist_ok=True)
sqlite_file = os.path.join(user_data_dir, "doc_management.db")

# -----------------------------
# FLASK APP
# -----------------------------
app = Flask(
    __name__,
    template_folder=os.path.join(base_path, "templates"),
    static_folder=os.path.join(base_path, "static"),
)
application = app  # for WSGI servers that look for 'application'
# TODO: replace with a secure random key in production
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(32))

# -----------------------------
# DB URIs
# -----------------------------
MYSQL_URI = os.environ.get(
    "DATABASE_URL",
    "mysql+pymysql://root:system@localhost/doc_management"
)  
SQLITE_URI = f"sqlite:///{sqlite_file}"
# -----------------------------
# N/A helper function
# -----------------------------
def get_or_create_na(model, **filters):
    obj = model.query.filter_by(**filters, name="N/A").first()
    if not obj:
        obj = model(name="N/A", **filters)
        db.session.add(obj)
        db.session.flush()
    return obj

def _norm_func(text: str) -> str:
    return " ".join((text or "").lower().split())

def norm_expr(column):
    return func.lower(func.replace(func.trim(column), " ", ""))

def norm_py(s):
    return "".join((s or "").lower().split())
# -----------------------------
# N/A case_insensitive function
# -----------------------------
from sqlalchemy import func

def get_or_create_case_insensitive(model, name_field, value, **filters):
    if not value:
        return None

    value_norm = value.strip()

    query = model.query.filter(
        func.lower(getattr(model, name_field)) == value_norm.lower()
    )

    for k, v in filters.items():
        query = query.filter(getattr(model, k) == v)

    obj = query.first()

    if obj:
        return obj

    obj = model(**{name_field: value_norm, **filters})
    db.session.add(obj)
    db.session.flush()  # important: get id without commit
    return obj

# -----------------------------
# DUPLICATE AVOID HELPER
# -----------------------------
import re
import unicodedata

def normalize_series_name(text: str) -> str:
    if not text:
        return ""

    # remove accents
    text = unicodedata.normalize('NFKD', text)
    text = "".join(c for c in text if not unicodedata.combining(c))

    # lowercase
    text = text.lower()

    # remove punctuation
    text = re.sub(r'[^\w\s]', '', text)

    # collapse spaces
    text = " ".join(text.split())

    return text
# -----------------------------
# SQLite schema helper
# -----------------------------
def _compile_sqlite_type(sa_col):
    """Best effort: compile SA column type for SQLite ALTER ADD COLUMN."""
    try:
        return sa_col.type.compile(dialect=sqlite_dialect.dialect())
    except Exception:
        return "TEXT"  # safe default


def ensure_sqlite_schema_matches_models22():
    """Non-destructive SQLite schema sync.

    - Create missing tables
    - Add missing columns (NULL default)
    Never drop or rename anything.
    """
    print("🔧 SQLite: ensuring schema (create missing tables / add missing columns)")
    inspector = inspect(db.engine)
    existing_tables = set(inspector.get_table_names())

    # Create missing tables
    for table in db.metadata.sorted_tables:
        if table.name not in existing_tables:
            print(f"  🆕 table → {table.name}")
            table.create(db.engine)

    # Add missing columns
    for table in db.metadata.sorted_tables:
        if table.name not in existing_tables:
            # just created; no need to add columns
            continue
        existing_cols = {c["name"] for c in inspector.get_columns(table.name)}
        for col in table.columns:
            if col.name in existing_cols:
                continue

            coltype = _compile_sqlite_type(col)
            nullable = "NULL"
            default_clause = ""
            if getattr(col, "server_default", None) and getattr(col.server_default, "arg", None):
                try:
                    default_clause = f" DEFAULT {str(col.server_default.arg)}"
                except Exception:
                    pass

            sql = (
                f'ALTER TABLE "{table.name}" '
                f'ADD COLUMN "{col.name}" {coltype} {nullable}{default_clause};'
            )
            print(f"  ➕ column → {table.name}.{col.name} ({coltype})")
            with db.engine.begin() as conn:
                conn.exec_driver_sql(sql)

# -----------------------------
# MySQL reachability helper
# -----------------------------
def mysql_reachable1() -> bool:
    """Check quickly if MySQL is reachable; used to decide DB backend."""
    try:
        conn = pymysql.connect(
            host="localhost",
            user="root",
            password="system",
            database="doc_management",
            connect_timeout=2,
        )
        conn.close()
        return True
    except Exception:
        return False

# -----------------------------
# Single DB configuration/init
# -----------------------------
def configure_database():

    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    # ============================================
    # SQLITE DATABASE
    # ============================================
    print("✅ Using SQLite")

    app.config["SQLALCHEMY_DATABASE_URI"] = (
        "sqlite:///doc_management.db"
    )

    # ============================================
    # SAFETY CHECK
    # ============================================
    if not app.config.get("SQLALCHEMY_DATABASE_URI"):
        raise RuntimeError("❌ DATABASE URI is not configured")

    # ============================================
    # INIT EXTENSIONS
    # ============================================
    from extensions import db, migrate

    db.init_app(app)
    migrate.init_app(app, db)   # ✅ REQUIRED FOR ALEMBIC

    # ============================================
    # VALIDATION (NO CREATE / NO ALTER)
    # ============================================
    with app.app_context():
        try:
            from sqlalchemy import text
            db.session.execute(text("SELECT 1"))
            print("✅ Database connection OK")

            # 🔥 ADD THIS BLOCK
            #if app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite"):
                #ensure_sqlite_schema_matches_models()

        except Exception as e:
            print("❌ Database connection failed:", e)
            raise

# Configure DB BEFORE importing models
configure_database()

# Now we can import models
from models import (
    Entity,
    Dependency,
    DependencyFunction,
    Series,
    Subseries,
    Doctype,
    DocumentalStudy
    )

# =========================================================
# ROUTES AND HELPERS
# (below this line the original route logic is preserved)
# =========================================================
@app.route('/')
def index():
    return render_template('index.html')
# =========================================================
# HELPER N/A NORMALIZER
# =========================================================
def normalize_na(value):
    if not value or str(value).strip().upper() == "N/A":
        return None
    return value
# =========================================================
# HELPER dependencies TREE
# =========================================================

def build_dependencies_hierarchy(entity_id=None):
    q = Dependency.query

    if entity_id:
        q = q.filter(Dependency.entity_id == entity_id)

    dependencies = q.order_by(Dependency.code.asc()).all()

    children_map = defaultdict(list)
    for d in dependencies:
        if d.parent_id:
            children_map[d.parent_id].append(d)

    result = []

    def walk(dep, level):
        result.append((dep, level))
        for child in children_map.get(dep.id, []):  # ✅ FIXED
            walk(child, level + 1)

    roots = [d for d in dependencies if not d.parent_id]

    for root in roots:
        walk(root, 0)

    return result

# =========================================================
# HELPER TVD STRUCTURE BUILDER
# =========================================================

def build_tvd_structure(studies, entity):

    from models import DependencyFunction

    functions_map = {
        f.id: f
        for f in DependencyFunction.query.all()
    }

    separator = entity.separator or "-"
    estructura = {}

    # =====================================
    # UNIQUE ORDERS
    # =====================================
    unique_series = []
    unique_subseries = {}

    for s in studies:
        if not s.serie:
            continue

        serie_name = s.serie.name.strip()
        sub_name = s.subserie.name.strip() if s.subserie else "N/A"

        if serie_name not in unique_series:
            unique_series.append(serie_name)

        unique_subseries.setdefault(serie_name, [])
        if sub_name not in unique_subseries[serie_name]:
            unique_subseries[serie_name].append(sub_name)

    # =====================================
    # MAIN LOOP
    # =====================================
    for s in studies:

        if not s.serie:
            continue

        func = functions_map.get(s.funciones_especificas)

        function_number = func.function_number if func else ""
        function_name = func.name if func else ""

        serie_key = s.serie.name.strip()
        sub_key = s.subserie.name.strip() if s.subserie else "N/A"

        serie_order = unique_series.index(serie_key) + 1
        sub_order = unique_subseries[serie_key].index(sub_key) + 1

        dep_code = (
            str(s.dependency.code).strip()
            if getattr(s, "dependency", None) and s.dependency.code
            else ""
        )

        # =====================================
        # DISTINCT CODES
        # =====================================
        serie_code = f"{dep_code}{separator}{serie_order}"
        sub_code = f"{serie_code}{separator}{sub_order}"
        # doc_code = f"{sub_code}{separator}{s.id}"

        # ================= SERIE =================
        if serie_key not in estructura:
            estructura[serie_key] = {
                "codigo": serie_code,
                "nombre": serie_key,
                "function_number": function_number,
                "function_name": function_name,
                "subseries": {}
            }

        # ================= SUBSERIE =================
        if sub_key not in estructura[serie_key]["subseries"]:
            estructura[serie_key]["subseries"][sub_key] = {
                "codigo": sub_code,
                "nombre": sub_key,
                "function_number": function_number,
                "function_name": function_name,
                "doctypes": []
            }
        existing_docs = estructura[serie_key]["subseries"][sub_key]["doctypes"]

        doc_order = len(existing_docs) + 1

        doc_code = (
            f"{sub_code}{separator}{doc_order}"
            if s.doctype else sub_code
        )

        # ================= DOCTYPE =================
        estructura[serie_key]["subseries"][sub_key]["doctypes"].append({
            "codigo": doc_code,
            "nombre": s.doctype.name if s.doctype else "N/A",
            "spt_fisico": s.spt_fisico,
            "spt_digital": s.spt_digital,
            "trd_annosOficina": s.trd_annosOficina,
            "trd_annosTransfAC": s.trd_annosTransfAC,
            "disposicion": getattr(s, "disposicion_final", ""),
            "reproduccion_tecnica": s.reproduccion_tecnica,
            "es_ddhh": s.prodDerechosHumanos,
            "function_number": function_number,
            "function_name": function_name,
            "procedimiento": getattr(s, "procedimiento", "")
        })

    result = []

    for serie in estructura.values():
        serie["subseries"] = list(serie["subseries"].values())
        result.append(serie)

    return result

# =========================================================
# ENTITY CRUD
# =========================================================
@app.route('/entities')
def list_entities():
    entities = Entity.query.order_by(Entity.name).all()

    return render_template(
        'list_entities.html',
        entities=entities,
        get_logo_url=get_logo_url
    )


import os
from flask import (
    request, redirect, url_for,
    render_template, flash, current_app
)
from werkzeug.utils import secure_filename
from models import Entity
from extensions import db

MAX_LOGO_URL_LENGTH = 1000
LOGO_CUT_LOG_FILE = "logo_url_cuts.log"

ALLOWED_IMAGE_MIME = (
    'image/png',
    'image/jpeg',
    'image/jpg',
    'image/svg+xml',
    'image/webp'
)

def _log_logo_cut(original, final):
    """Log URL cuts for auditing/debugging."""
    with open(LOGO_CUT_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(
            f"[{datetime.now().isoformat()}] "
            f"CUT LOGO URL\n"
            f"  original: {original}\n"
            f"  saved:    {final}\n\n"
        )


def normalize_logo_url(url, max_len=MAX_LOGO_URL_LENGTH):
    """
    VALIDATE → CUT → LOG

    - Validates URL structure
    - Converts Google Drive share links to direct image links
    - Verifies image MIME via HEAD
    - Cuts URL ONLY if it exceeds max_len
    - Logs whenever a cut occurs
    """
    if not url:
        return None

    url = url.strip()

    # -------------------------------
    # Validate URL structure
    # -------------------------------
    parsed = urlparse(url)
    if not parsed.scheme or not parsed.netloc:
        return None

    # -------------------------------
    # Google Drive share → direct
    # -------------------------------
    if "drive.google.com" in url and "/file/d/" in url:
        try:
            file_id = url.split("/file/d/")[1].split("/")[0]
            url = f"https://drive.google.com/uc?id={file_id}"
            parsed = urlparse(url)
        except IndexError:
            return None

    # -------------------------------
    # Validate MIME type (HEAD)
    # -------------------------------
    # try:
    #    req = Request(url, method="HEAD")
    #    with urlopen(req, timeout=5) as resp:
    #        content_type = resp.headers.get("Content-Type", "")
    #    if not content_type.startswith("image/"):
    #        return None
    # except (HTTPError, URLError):
    #    return None
    # 🔥 RELAXED VALIDATION
    # Only basic URL validation
    # return url
    # -------------------------------
    # CUT if needed (safe-first)
    # -------------------------------
    original_url = url

    if len(url) > max_len:
        # Try removing query + fragment first
        safe_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"

        if len(safe_url) <= max_len:
            url = safe_url
        else:
            # Last-resort hard cut
            url = url[:max_len]

        _log_logo_cut(original_url, url)

    return url

from urllib.parse import urlparse
from flask import request, redirect, url_for, flash, render_template

def get_logo_url(entity):
    logo = (entity.logo_path or "").strip()

    if not logo:
        return url_for('static', filename='logos/default.png')

    if logo.startswith(("http", "data:image")):
        return logo

    return url_for('static', filename=logo)


from urllib.parse import urlparse
import base64

# ==========================================================
# ADD ENTITY
# ==========================================================
@app.route('/entities/add', methods=['POST'])
def add_entity():

    raw_logo = request.form.get('logo_path', '').strip()
    logo_value = None

    if raw_logo:

        # CASE 1: pasted image (base64)
        if raw_logo.startswith('data:image/'):
            logo_value = raw_logo

        # CASE 2: normal URL
        else:
            parsed = urlparse(raw_logo)

            if parsed.scheme in ('http', 'https') and parsed.netloc:
                logo_value = normalize_logo_url(raw_logo)
            else:
                flash("Invalid logo format.", "danger")
                return redirect(url_for('list_entities'))

    entity = Entity(
        name=request.form['name'],
        short_name=request.form.get('short_name'),
        nit=request.form.get('nit'),
        email=request.form.get('email'),
        phone=request.form.get('phone'),
        address=request.form.get('address'),
        acto_administrativo=request.form.get('acto_administrativo'),
        functions=request.form.get('functions'),
        separator=request.form.get('separator'),
        central_dependency_code=request.form.get('central_dependency_code'),
        is_centralized=('is_centralized' in request.form),
        logo_path=logo_value,
        is_active=('is_active' in request.form),

        representante_legal_nombre=request.form.get('representante_legal_nombre'),
        representante_legal_cargo=request.form.get('representante_legal_cargo'),
        nom_responsable_gestion_documental=request.form.get('nom_responsable_gestion_documental'),
        cgo_responsable_gestion_documental=request.form.get('cgo_responsable_gestion_documental'),
        dep_responsable_gestion_documental=request.form.get('dep_responsable_gestion_documental')
    )

    db.session.add(entity)
    db.session.commit()

    flash("Entity added successfully!", "success")
    return redirect(url_for('list_entities'))


# ==========================================================
# EDIT ENTITY
# ==========================================================
@app.route('/entities/edit/<int:id>', methods=['GET', 'POST'])
@app.route('/edit_entity/<int:id>', methods=['GET', 'POST'])
def edit_entity(id):

    entity = Entity.query.get_or_404(id)

    if request.method == 'POST':

        entity.name = request.form.get('name')
        entity.short_name = request.form.get('short_name')
        entity.nit = request.form.get('nit')
        entity.email = request.form.get('email')
        entity.phone = request.form.get('phone')
        entity.address = request.form.get('address')
        entity.acto_administrativo = request.form.get('acto_administrativo')
        entity.separator = request.form.get('separator')
        entity.central_dependency_code = request.form.get('central_dependency_code')
        entity.functions = request.form.get('functions')

        entity.representante_legal_nombre = request.form.get('representante_legal_nombre')
        entity.representante_legal_cargo = request.form.get('representante_legal_cargo')
        entity.nom_responsable_gestion_documental = request.form.get('nom_responsable_gestion_documental')
        entity.cgo_responsable_gestion_documental = request.form.get('cgo_responsable_gestion_documental')
        entity.dep_responsable_gestion_documental = request.form.get('dep_responsable_gestion_documental')

        # =====================================
        # LOGO SUPPORTS:
        # 1. BASE64 pasted image
        # 2. HTTP URL
        # 3. empty = remove
        # =====================================
        raw_logo = request.form.get('logo_path', '').strip()

        if raw_logo:

            # pasted image
            if raw_logo.startswith('data:image/'):
                entity.logo_path = raw_logo

            # url image
            else:
                parsed = urlparse(raw_logo)

                is_valid_scheme = parsed.scheme in ('http', 'https') and parsed.netloc
                is_valid_ext = parsed.path.lower().endswith(
                    ('.jpg', '.jpeg', '.png', '.webp', '.svg')
                )

                if is_valid_scheme and is_valid_ext:
                    entity.logo_path = normalize_logo_url(raw_logo)
                else:
                    flash("Invalid logo.", "danger")
                    return redirect(url_for('list_entities'))

        else:
            entity.logo_path = None

        # checkboxes
        entity.is_centralized = 'is_centralized' in request.form
        entity.is_active = 'is_active' in request.form

        try:
            db.session.commit()
            flash("Entidad actualizada correctamente.", "success")

        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")

        return redirect(url_for('list_entities'))

    return render_template('edit_entities.html', entity=entity)

@app.route('/entities/delete/<int:id>')
def delete_entity(id):
    entity = Entity.query.get_or_404(id)
    db.session.delete(entity)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")
    
    flash("Entity deleted successfully!", "success")
    return redirect(url_for('list_entities'))


@app.route("/select_entity", methods=["POST"])
def select_entity():
    eid = request.form.get("entity_id", type=int)
    if not eid or not Entity.query.get(eid):
        flash("Entidad inválida.", "danger")
        return redirect(request.referrer or url_for("list_entities"))
    session["active_entity_id"] = eid
    session.modified = True
    flash("Entidad seleccionada actualizada.", "success")
    return redirect(request.referrer or url_for("list_entities"))


@app.context_processor
def inject_active_entity():
    eid = session.get("active_entity_id")
    return {
        "active_entity_id": eid,
        "active_entity": Entity.query.get(eid) if eid else None
    }
    
@app.get("/api/entities")
def api_get_entities():
    try:
        entities = Entity.query.order_by(Entity.name).all()

        return jsonify([
            {
                "id": e.id,
                "name": e.name,
                "acto_administrativo": e.acto_administrativo or ""
            }
            for e in entities
        ])

    except Exception as e:
        print("❌ ERROR /api/entities:", e)
        return jsonify([]), 500

@app.route("/add_entity_inline", methods=["POST"])

def add_entity_inline():

    data = request.get_json() or {}

    raw_logo = (data.get("logo_path") or "").strip()
    logo_value = None

    if raw_logo:

        # CASE 1: pasted image (base64)
        if raw_logo.startswith("data:image/"):
            logo_value = raw_logo

        # CASE 2: URL
        else:
            parsed = urlparse(raw_logo)

            if parsed.scheme in ("http", "https") and parsed.netloc:
                logo_value = normalize_logo_url(raw_logo)
            else:
                return jsonify({"error": "Invalid logo format"}), 400

    entity = Entity(
        name=data.get("name"),
        short_name=data.get("short_name"),
        nit=data.get("nit"),
        email=data.get("email"),
        phone=data.get("phone"),
        address=data.get("address"),
        acto_administrativo=data.get("acto_administrativo"),
        separator=data.get("separator"),
        logo_path=logo_value,   # ✅ THIS WAS MISSING
        is_active=True
    )

    db.session.add(entity)
    db.session.commit()

    return jsonify({"id": entity.id})
# =========================================================
# dependencies
# =========================================================

@app.route('/dependencies')
def list_dependencies():

    active_entity_id = request.args.get("entity_id", type=int) \
        or session.get("active_entity_id")

    # 🔥 update session if changed from dropdown
    if active_entity_id:
        session["active_entity_id"] = active_entity_id

    entities = Entity.query.order_by(Entity.name).all()

    q = Dependency.query

    if active_entity_id:
        q = q.filter(Dependency.entity_id == active_entity_id)

    dependencies = q.order_by(Dependency.code.asc()).all()

    return render_template(
        'list_dependencies.html',
        dependencies=dependencies,
        dependencies_hierarchy=build_dependencies_hierarchy(active_entity_id),
        entities=entities,  # ✅ NEW
        active_entity_id=active_entity_id  # ✅ ensure passed
    )


import re;
# =========================================================
# ADD dependencies WITH SMART FUNCTIONS PARSING (FIXED)
# =========================================================

@app.route('/dependencies/add', methods=['POST'])
def add_dependencies():

    from sqlalchemy import func
    from sqlalchemy.exc import IntegrityError
    import re

    name = request.form.get('name', '').strip()
    code = request.form.get('code', '').strip()
    entity_id = request.form.get('entity_id', type=int)
    dependency_of = request.form.get('dependency_of', type=int)
    functions_text = request.form.get('functions', '').strip()
    boss_name = request.form.get('boss_name', '').strip()
    boss_charge = request.form.get('boss_charge', '').strip()
    legal_representante = request.form.get('legal_representante') == '1'
    tvd_responsible = request.form.get('tvd_responsible') == '1'

    if not name:
        flash("El nombre de la Dependencia es obligatorio.", "warning")
        return redirect(url_for('list_dependencies'))

    if not entity_id:
        flash("Debe seleccionar una Entidad.", "warning")
        return redirect(url_for('list_dependencies'))

    if dependency_of:
        parent = db.session.get(Dependency, dependency_of)
        if not parent or parent.entity_id != entity_id:
            flash("Dependencia padre inválida.", "danger")
            return redirect(url_for('list_dependencies'))

    exists = db.session.query(Dependency).filter(
        func.lower(Dependency.name) == name.lower(),
        Dependency.entity_id == entity_id
    ).first()

    if exists:
        flash("Ya existe una Dependencia con ese nombre.", "warning")
        return redirect(url_for('list_dependencies'))

    try:
        dep = Dependency(
            name=name,
            code=code,
            entity_id=entity_id,
            parent_id=dependency_of,
            boss_name=boss_name,
            boss_charge=boss_charge,
            tvd_responsible=tvd_responsible,
            legal_representante=legal_representante
        )

        db.session.add(dep)
        db.session.flush()

        # ✅ FUNCTIONS PARSING (single block)
        pattern = r'(\d+)\.\s*(.+?)(?=\n\d+\.|\Z)'
        matches = re.findall(pattern, functions_text, re.DOTALL)

        seen = set()

        if matches:
            for number, text in matches:
                num = int(number)
                if num in seen:
                    continue
                seen.add(num)

                db.session.add(DependencyFunction(
                    dependency_id=dep.id,
                    function_number=num,
                    name=" ".join(text.split()),
                    is_non_documented=False
                ))
        else:
            lines = [l.strip() for l in functions_text.split("\n") if l.strip()]
            for i, line in enumerate(lines, start=1):
                db.session.add(DependencyFunction(
                    dependency_id=dep.id,
                    function_number=i,
                    name=line,
                    is_non_documented=False
                ))

        db.session.commit()

    except IntegrityError:
        db.session.rollback()
        flash("❌ No se permiten números de función duplicados.", "danger")
        return redirect(url_for('list_dependencies'))

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error inesperado: {str(e)}", "danger")
        return redirect(url_for('list_dependencies'))

    flash("Dependency added successfully!", "success")
    return redirect(url_for('list_dependencies'))

# =========================================================
# EDIT dependencies (FIXED & SAFE)
# =========================================================

@app.route('/edit_dependencies/<int:id>', methods=['GET', 'POST'])
def edit_dependencies(id):

    from sqlalchemy import func
    from sqlalchemy.exc import IntegrityError

    dep = db.session.get(Dependency, id)
    if not dep:
        abort(404)

    dependencies_hierarchy = build_dependencies_hierarchy(dep.entity_id) or []

    if request.method == 'POST':

        # ===============================
        # GET & CLEAN DATA
        # ===============================
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip()
        dependency_of = request.form.get('dependency_of', type=int)
        tvd_responsible = '1' in request.form.getlist('tvd_responsible')
        legal_representante = '1' in request.form.getlist('legal_representante')
        boss_name = request.form.get('boss_name', '').strip()
        boss_charge = request.form.get('boss_charge', '').strip()

        # ===============================
        # VALIDATIONS
        # ===============================
        if not name:
            flash("El nombre de la Dependencia es obligatorio.", "warning")
            return redirect(url_for('edit_dependencies', id=id))

        # ---------- PREVENT SELF-PARENT ----------
        if dependency_of == dep.id:
            flash("Una dependencia no puede ser su propio padre.", "danger")
            return redirect(url_for('edit_dependencies', id=id))

        # ---------- VALIDATE PARENT ----------
        if dependency_of:
            parent = db.session.get(Dependency, dependency_of)

            if not parent or parent.entity_id != dep.entity_id:
                flash("Dependencia padre inválida.", "danger")
                return redirect(url_for('edit_dependencies', id=id))

        # ---------- AVOID DUPLICATES ----------
        exists = db.session.query(Dependency).filter(
            func.lower(Dependency.name) == name.lower(),
            Dependency.entity_id == dep.entity_id,
            Dependency.id != dep.id  # exclude itself
        ).first()

        if exists:
            flash("Ya existe una Dependencia con ese nombre.", "warning")
            return redirect(url_for('edit_dependencies', id=id))

        # ===============================
        # UPDATE
        # ===============================
        try:
            # ===============================
            # UPDATE FIELDS
            # ===============================
            dep.name = name
            dep.code = code
            dep.parent_id = dependency_of
            dep.tvd_responsible = tvd_responsible
            dep.legal_representante = legal_representante
            dep.boss_name = boss_name
            dep.boss_charge = boss_charge

            # ===============================
            # ENFORCE UNIQUE FLAGS
            # ===============================
            if tvd_responsible:
                db.session.query(Dependency).filter(
                    Dependency.entity_id == dep.entity_id,
                    Dependency.id != dep.id,
                    Dependency.tvd_responsible == True
                ).update({"tvd_responsible": False}, synchronize_session=False)

            if legal_representante:
                db.session.query(Dependency).filter(
                    Dependency.entity_id == dep.entity_id,
                    Dependency.id != dep.id,
                    Dependency.legal_representante == True
                ).update({"legal_representante": False}, synchronize_session=False)

            # ===============================
            # SINGLE COMMIT (ONLY HERE)
            # ===============================
            db.session.commit()

        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error: {str(e)}", "danger")
            return redirect(url_for('edit_dependencies', id=id))

        flash('Dependency updated successfully.', 'success')
        return redirect(url_for('list_dependencies'))

    return render_template(
        'edit_dependencies.html',
        dep=dep,
        dependencies_hierarchy=dependencies_hierarchy
    )

#========================================================
# FUNCTIONS PREVIEW (AJAX)
#========================================================

@app.route('/dependencies/preview_functions', methods=['POST'])
def preview_functions():
    import re

    functions_text = request.form.get('functions', '')

    pattern = r'(\d+)\.\s*(.+?)(?=\n\d+\.|\Z)'
    matches = re.findall(pattern, functions_text, re.DOTALL)

    result = []

    # CASE 1: numbered input
    if matches:
        for number, text in matches:
            clean_text = " ".join(text.split())
            result.append({
                "number": int(number),
                "text": clean_text
            })

    # CASE 2: fallback
    else:
        lines = [l.strip() for l in functions_text.split("\n") if l.strip()]
        for i, line in enumerate(lines, start=1):
            result.append({
                "number": i,
                "text": line
            })

    return jsonify(result)
# =========================================================
# DELETE dependencies (with cascade delete of functions)
# =========================================================
@app.route('/dependencies/delete/<int:id>')
def delete_dependencies(id):
    dep = Dependency.query.get_or_404(id)
    db.session.delete(dep)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("Error deleting dependency!", "danger")
    return redirect(url_for('list_dependencies'))

# =========================================================
# dependencies FUNCTIONS
# =========================================================
# =========================================================
# API DEPENDENCIES (FIXED)
# =========================================================
@app.get("/api/dependencies")
def get_dependencies():

    entity_id = request.args.get("entity_id", type=int)

    if not entity_id:
        return jsonify([])

    try:
        dependencies = (
            Dependency.query
            .filter(Dependency.entity_id == entity_id)
            .order_by(Dependency.name)
            .all()
        )

        return jsonify([
            {
                "id": d.id,
                "name": d.name,
                "boss_name": d.boss_name or ""
            }
            for d in dependencies
        ])

    except Exception as e:
        print("❌ ERROR /api/dependencies:", e)
        return jsonify([]), 500
    
@app.route('/add_dependencies_inline', methods=['POST'])
def add_dependencies_inline():

    data = request.get_json() or {}

    entity_id = data.get("entity_id")
    code = data.get("code")
    name = data.get("name")
    dependency_of = data.get("dependency_of")
    boss_name = (data.get("boss_name") or "").strip()
    boss_charge = (data.get("boss_charge") or "").strip()
    tvd_responsible = data.get("tvd_responsible") is True
    legal_representante = data.get("legal_representante") is True
    functions_text = (data.get("functions") or "").strip()

    if not entity_id or not name:
        return jsonify({"error": "Entidad y nombre son obligatorios"}), 400

    dep = Dependency(
        entity_id=entity_id,
        code=code,
        name=name,
        parent_id=dependency_of,
        boss_name=boss_name,
        boss_charge=boss_charge,
        tvd_responsible=tvd_responsible,
        legal_representante=legal_representante
    )

    db.session.add(dep)
    db.session.flush()

    # ✅ FUNCTIONS PARSING
    import re
    pattern = r'(\d+)\.\s*(.+?)(?=\n\d+\.|\Z)'
    matches = re.findall(pattern, functions_text, re.DOTALL)

    seen = set()

    if matches:
        for number, text in matches:
            num = int(number)
            if num in seen:
                continue
            seen.add(num)

            db.session.add(DependencyFunction(
                dependency_id=dep.id,
                function_number=num,
                name=" ".join(text.split()),
                is_non_documented=False
            ))
    else:
        lines = [l.strip() for l in functions_text.split("\n") if l.strip()]
        for i, line in enumerate(lines, start=1):
            db.session.add(DependencyFunction(
                dependency_id=dep.id,
                function_number=i,
                name=line,
                is_non_documented=False
            ))

    db.session.commit()

    return jsonify({
        "id": dep.id,
        "name": dep.name
    }), 201

@app.route('/set_function_non_documented/<int:function_id>', methods=['POST'])
def set_function_non_documented(function_id):
    df = DependencyFunction.query.get(function_id)
    if not df:
        return jsonify({'success': False, 'error': 'Function not found'}), 404

    data = request.get_json(silent=True) or {}
    #bool(request.form.get("field"))    
    value = data.get("is_non_documented", False)

    df.is_non_documented = value
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return jsonify({
        'success': True,
        'non_documented': df.is_non_documented
    })
# =========================================================
# SERIES CRUD
# =========================================================
@app.route('/series')
def list_series():
    series_list = Series.query.order_by(Series.name).all()
    return render_template('list_series.html', series=series_list)


@app.route('/add_series', methods=['POST'])
def add_series():
    name = request.form['name'].strip()
    definition = request.form['definition'].strip()

    support = request.form.get('support')
    retention_management = request.form.get('retention_management', type=int)
    retention_central = request.form.get('retention_central', type=int)
    final_disposition = request.form.get('final_disposition')
    notes = request.form.get('notes')

    # ================= VALIDATION =================
    if not name or not definition:
        flash("Datos inválidos.", "warning")
        return redirect(url_for('list_series'))

    # 🔥 TRD VALIDATION (IMPORTANT)
    if retention_management and retention_central:
        if retention_central < retention_management:
            flash("La retención en archivo central no puede ser menor que en gestión.", "warning")
            return redirect(url_for('list_series'))

    # ================= DUPLICATE CHECK =================
    normalized_input = normalize_series_name(name)

    all_series = Series.query.all()

    exists = next(
        (s for s in all_series if normalize_series_name(s.name) == normalized_input),
        None
    )

    if exists:
        flash("Ya existe una Serie con ese nombre.", "warning")
        return redirect(url_for('list_series'))

    # ================= CREATE =================
    serie = Series(
        name=name,
        definition=definition,
        support=support,
        retention_management=retention_management,
        retention_central=retention_central,
        final_disposition=final_disposition,
        notes=notes
    )
    VALID_DISPOSITIONS = {"Eliminación", "Selección", "Conservación total"}

    if final_disposition and final_disposition not in VALID_DISPOSITIONS:
        flash("Disposición final inválida.", "warning")
        return redirect(url_for('list_series'))

    db.session.add(serie)
    db.session.commit()

    flash("Series added successfully!", "success")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({
            "id": serie.id,
            "name": serie.name
        })
    return redirect(url_for('list_series'))

@app.route('/series/edit/<int:id>', methods=['GET', 'POST'])
def edit_series(id):
    serie = Series.query.get_or_404(id)

    if request.method == 'POST':
        new_name = request.form.get('name', '').strip()
        new_definition = request.form.get('definition', '').strip()

        support = request.form.get('support')
        retention_management = request.form.get('retention_management', type=int)
        retention_central = request.form.get('retention_central', type=int)
        final_disposition = request.form.get('final_disposition')
        notes = request.form.get('notes')

        # ================= VALIDATION =================
        if not new_name or not new_definition:
            flash("Datos inválidos.", "warning")
            return redirect(url_for('edit_series', id=id))

        # 🔥 TRD VALIDATION
        if retention_management and retention_central:
            if retention_central < retention_management:
                flash("La retención en archivo central no puede ser menor que en gestión.", "warning")
                return redirect(url_for('edit_series', id=id))

        # 🔥 DUPLICATE CHECK (excluding current)
        normalized_input = normalize_series_name(new_name)

        all_series = Series.query.filter(Series.id != id).all()

        exists = next(
            (s for s in all_series if normalize_series_name(s.name) == normalized_input),
            None
        )

        if exists:
            flash("Ya existe otra Serie con ese nombre.", "warning")
            return redirect(url_for('edit_series', id=id))

        # ================= UPDATE =================
        serie.name = new_name
        serie.definition = new_definition
        serie.support = support
        serie.retention_management = retention_management
        serie.retention_central = retention_central
        serie.final_disposition = final_disposition
        serie.notes = notes

        db.session.commit()

        flash("Series updated successfully!", "success")
        return redirect(url_for('list_series'))

    return render_template('edit_series.html', serie=serie)

@app.route('/series/delete/<int:id>')
def delete_series(id):
    serie = Series.query.get_or_404(id)

    if serie.subseries:
        flash("Cannot delete series with subseries.", "danger")
        return redirect(url_for('list_series'))
    
    db.session.delete(serie)
    db.session.commit()
    return redirect(url_for('list_series'))

# =========================================================
# SUBSERIES CRUD
# =========================================================
@app.route('/subseries')

def list_subseries():
    subseries_list = Subseries.query.all()
    series_list = Series.query.order_by(Series.name).all()
    return render_template('list_subseries.html', subseries=subseries_list, series=series_list)

@app.route('/subseries/add', methods=['POST'])
def add_subseries():
    name = request.form.get('name', '').strip()
    definition = request.form.get('definition', '').strip()
    series_id = request.form.get('series_id', type=int)

    support = request.form.get('support')
    retention_management = request.form.get('retention_management', type=int)
    retention_central = request.form.get('retention_central', type=int)
    final_disposition = request.form.get('final_disposition')
    notes = request.form.get('notes')

    # ================= VALIDATION =================
    if not name or not definition or not series_id:
        flash("Datos inválidos.", "warning")
        return redirect(url_for('list_subseries'))

    # 🔥 TRD VALIDATION
    if retention_management and retention_central:
        if retention_central < retention_management:
            flash("La retención en archivo central no puede ser menor que en gestión.", "warning")
            return redirect(url_for('list_subseries'))

    # 🔥 FINAL DISPOSITION VALIDATION
    VALID_DISPOSITIONS = {"Eliminación", "Selección", "Conservación total"}
    if final_disposition and final_disposition not in VALID_DISPOSITIONS:
        flash("Disposición final inválida.", "warning")
        return redirect(url_for('list_subseries'))

    # ================= SERIES CHECK =================
    serie = Series.query.get(series_id)
    if not serie:
        flash("Serie no encontrada.", "warning")
        return redirect(url_for('list_subseries'))

    # ================= DUPLICATE CHECK =================
    exists = Subseries.query.filter(
        func.lower(Subseries.name) == name.lower(),
        Subseries.series_id == series_id
    ).first()

    if exists:
        flash("Ya existe una Subserie con ese nombre en la serie.", "warning")
        return redirect(url_for('list_subseries'))

    # ================= CREATE =================
    subserie = Subseries(
        name=name,
        definition=definition,
        series_id=series_id,
        support=support,
        retention_management=retention_management,
        retention_central=retention_central,
        final_disposition=final_disposition,
        notes=notes
    )

    db.session.add(subserie)
    db.session.commit()

    flash("Subseries added successfully!", "success")
    return redirect(url_for('list_subseries'))

@app.route('/subseries/edit/<int:id>', methods=['GET', 'POST'])
def edit_subseries(id):
    subserie = Subseries.query.get_or_404(id)

    if request.method == 'POST':
        new_name = request.form.get('name', '').strip()
        new_definition = request.form.get('definition', '').strip()
        series_id = request.form.get('series_id', type=int)

        if not new_name or not new_definition or not series_id:
            flash("Datos inválidos.", "warning")
            return redirect(url_for('list_subseries'))

        subserie.name = new_name
        subserie.definition = new_definition
        subserie.series_id = series_id

        subserie.support = request.form.get('support')
        subserie.retention_management = request.form.get('retention_management', type=int)
        subserie.retention_central = request.form.get('retention_central', type=int)
        subserie.final_disposition = request.form.get('final_disposition')
        subserie.notes = request.form.get('notes')

        db.session.commit()

        flash("Subseries updated successfully!", "success")
        return redirect(url_for('list_subseries'))

    series_list = Series.query.order_by(Series.name).all()

    return render_template(
        'edit_subseries.html',
        subseries=subserie,
        series=series_list
    )

@app.route('/subseries/delete/<int:id>')
def delete_subseries(id):
    subserie = Subseries.query.get_or_404(id)
    db.session.delete(subserie)
    db.session.commit()
    flash("Subseries deleted successfully!", "success")
    return redirect(url_for('list_subseries'))

# =========================================================
# DOCTYPES CRUD
# =========================================================
@app.route('/doctypes')
def list_doctypes():
    doctypes = (
        Doctype.query
        .options(
            joinedload(Doctype.subseries)
            .joinedload(Subseries.series)
        )
        .all()
    )

    subseries = Subseries.query.order_by(Subseries.name).all()

    return render_template(
        'list_doctypes.html',
        doctypes=doctypes,
        subseries=subseries
    )

@app.route('/add_doctype', methods=['POST'])
def add_doctype():
    name = request.form['name'].strip()
    definition = request.form['definition'].strip()
    subserie_id = request.form['subserie_id']
    subserie = Subseries.query.get(subserie_id)

    if not subserie:
        flash("Subserie no encontrada.", "warning")
        return redirect(url_for('list_doctypes'))

    exists = Doctype.query.filter(
        func.lower(Doctype.name) == name.lower(),
        Doctype.subseries_id == subserie.id
    ).first()
    if exists:
        flash("Ya existe un Tipo Documental con ese nombre en la Subserie seleccionada.", "warning")
        return redirect(url_for('list_doctypes'))

    doctype = Doctype(name=name, subseries_id=subserie.id)
    db.session.add(doctype)
    db.session.commit()
    flash("Tipo Documental added successfully!", "success")
    return redirect(url_for('list_doctypes'))


@app.route('/doctypes/edit/<int:id>', methods=['GET', 'POST'])
def edit_doctype(id):
    doctype = Doctype.query.get_or_404(id)
    if request.method == 'POST':
        new_name = request.form['name'].strip()
        doctype.name = new_name
        doctype.definition = request.form.get('definition', '').strip()
        db.session.commit()
        flash("Tipo Documental updated successfully!", "success")
        return redirect(url_for('list_doctypes'))
    return render_template('edit_doctype.html', doctype=doctype)


@app.route('/doctypes/delete/<int:id>')
def delete_doctype(id):
    doctype = Doctype.query.get_or_404(id)
    db.session.delete(doctype)
    db.session.commit()
    flash("Tipo Documental deleted successfully!", "success")
    return redirect(url_for('list_doctypes'))

# =========================================================
# AJAX ADD DOCTYPE
# =========================================================
@app.route('/add_doctype_ajax', methods=['POST'])
def add_doctype_ajax():
    name = request.form.get('name', '').strip()
    definition = request.form.get('definition', '').strip() or "N/A"
    if not name:
        return jsonify({'error': 'Nombre requerido'}), 400

    subserie_id = session.get('last_subserie_id')
    if not subserie_id:
        last_subserie = Subseries.query.order_by(Subseries.series_id.desc()).first()
        subserie_id = last_subserie.id if last_subserie else None

    if not subserie_id:
        return jsonify({'error': 'No existe una subserie para asociar'}), 400

    existing = Doctype.query.filter(
        func.lower(Doctype.name) == name.lower(),
        Doctype.subseries_id == subserie_id
    ).first()
    if existing:
        return jsonify({'error': 'Ya existe un tipo documental con ese nombre en esa subserie.'}), 400

    doctype = Doctype(
        name=name,
        definition=definition,
        subseries_id=subserie_id
    )
    db.session.add(doctype)
    db.session.commit()
    return jsonify({'id': doctype.id, 'name': doctype.name})


# =========================================================
# NEW REST API — Series/Subseries/Doctype for the new template
# =========================================================
@app.get("/api/series")
def api_get_series():
    name = request.args.get("name")
    series_id = request.args.get("id", type=int)

    if series_id:

        s = Series.query.get(series_id)

        if not s:
            return jsonify({})

        return jsonify({
            "id": s.id,
            "name": s.name,
            "final_disposition": s.final_disposition or ""
        })

        def normalize_disposition(value):
            return (
                "Conservación" if value in ["C", "Conservación total", "Conservación"] else
                "Selección" if value in ["S", "Selección"] else
                "Eliminación" if value in ["E", "Eliminación"] else
                ""
            )

    # =========================
    # SINGLE SERIE
    # =========================
    if name:
        s = Series.query.filter(func.upper(Series.name) == name.upper()).first()

        if not s:
            return jsonify({})

        return jsonify({
            "id": s.id,
            "name": s.name,
            "code": s.code or "",
            "retention": s.retention_central,
            "disposition": s.final_disposition,
            "definition": s.definition or ""
        })

    # =========================
    # ALL SERIES
    # =========================
    items = Series.query.order_by(Series.name).all()

    return jsonify([
        {
            "id": s.id,
            "name": s.name,
            "retention": s.retention_central,
            "disposition": s.final_disposition,
            "definition": s.definition or ""
        }
        for s in items
    ])

@app.post("/api/series")
def api_post_series():
    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()

    if not name:
        return jsonify({"error": "Nombre requerido"}), 400

    existing = Series.query.filter(
        func.lower(Series.name) == name.lower()
    ).first()

    if existing:
        return jsonify({
            "id": existing.id,
            "name": existing.name
        }), 200

    serie = Series(name=name)
    db.session.add(serie)
    db.session.commit()

    # ✅ FIXED RESPONSE
    return jsonify({
        "id": serie.id,
        "name": serie.name
    }), 201


@app.get("/api/subseries")
def api_subseries():
    series_id = request.args.get("series_id", type=int)

    if not series_id:
        return jsonify([])

    subseries = Subseries.query.filter_by(series_id=series_id)\
        .order_by(Subseries.name).all()

    return jsonify([
        {"id": s.id, "name": s.name}
        for s in subseries
    ])

@app.post("/api/subseries")
def api_post_subseries():
    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    series_id = data.get("series_id")
    definition = (data.get("definition") or "N/A").strip()

    if not name or not series_id:
        return jsonify({"error": "Nombre y serie son requeridos"}), 400

    existing = Subseries.query.filter(
        func.lower(Subseries.name) == name.lower(),
        Subseries.series_id == series_id
    ).first()

    if existing:
        return jsonify({
            "id": existing.id,
            "name": existing.name
        }), 200

    subserie = Subseries(
        name=name,
        series_id=series_id,
        definition=definition
    )

    db.session.add(subserie)
    db.session.commit()

    return jsonify({
        "id": subserie.id,
        "name": subserie.name
    }), 201

@app.post("/add_subseries_inline")
def add_subseries_inline():
    data = request.get_json() or {}

    name = (data.get("name") or "").strip()
    series_id = data.get("series_id")

    if not name or not series_id:
        return jsonify({"error": "Invalid data"}), 400

    # ✅ FIX: ALWAYS provide definition
    sub = get_or_create_case_insensitive(
        Subseries,
        "name",
        name,
        series_id=series_id,
        definition="N/A"   # 🔥 REQUIRED
    )

    db.session.commit()

    return jsonify({
        "id": sub.id,
        "name": sub.name
    })

    db.session.commit()

    return jsonify({
        "id": sub.id,
        "name": sub.name
    })


@app.get("/api/doctypes")
def api_get_doctypes():
    subserie_id = request.args.get("subserie_id", type=int)

    if not subserie_id:
        return jsonify([])

    doctypes = Doctype.query.filter_by(subseries_id=subserie_id).all()

    return jsonify([
        {"id": d.id, "name": d.name}
        for d in doctypes
    ])

@app.post("/api/doctypes")
def api_post_doctypes():
    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    subserie_id = data.get("subserie_id")
    definition = data.get("definition", "").strip()

    if not name or not subserie_id or not definition:
        return jsonify({"error": "Nombre, definición y Subserie son requeridos"}), 400

    existing = Doctype.query.filter(
        Doctype.name == name,
        Doctype.subseries_id == subserie_id
    ).first()

    if existing:
        return jsonify({
            "id": existing.id,
            "name": existing.name
        }), 200

    doctype = Doctype(
        name=name,
        subseries_id=subserie_id,
        definition=definition   
    )

    db.session.add(doctype)
    db.session.commit()

    return jsonify({
        "id": doctype.id,
        "name": doctype.name
    }), 201


@app.post("/add_doctype_inline")
def add_doctype_inline():
    data = request.get_json()

    name = (data.get("name") or "").strip()
    subseries_id = data.get("subseries_id")

    if not name or not subseries_id:
        return jsonify({"error": "Invalid data"}), 400

    doc = get_or_create_case_insensitive(
        Doctype,
        "name",
        name,
        subseries_id=subseries_id
    )

    db.session.commit()

    return jsonify({
        "id": doc.id,
        "name": doc.name
    })

# =========================================================
# DOCUMENTAL STUDY CRUD
# =========================================================
from collections import defaultdict

@app.route('/documental_studies')
def list_documental_studies():

    # ============================================
    # ENTITY SELECTION (NO SESSION DEPENDENCY)
    # ============================================
    active_entity_id = request.args.get("entity_id", type=int)

    if not active_entity_id:
        first = Entity.query.order_by(Entity.name).first()
        active_entity_id = first.id if first else None

    if active_entity_id:
        session["active_entity_id"] = active_entity_id

    # ============================================
    # ENTITIES FOR DROPDOWN
    # ============================================
    entities = Entity.query.order_by(Entity.name).all()

    # ============================================
    # STUDIES
    # ============================================
    studies = []
    if active_entity_id:
        studies = (
            DocumentalStudy.query
            .filter(DocumentalStudy.entity_id == active_entity_id)
            .all()
        )

    # ============================================
    # GROUP BY dependencies
    # ============================================
    from collections import defaultdict
    studies_by_dep = defaultdict(list)

    for s in studies:
        if s.dependency_id:
            studies_by_dep[s.dependency_id].append(s)

    # ============================================
    # dependencies HIERARCHY
    # ============================================
    dependencies_hierarchy = build_dependencies_hierarchy(active_entity_id)

    # ============================================
    # FUNCTIONS MAP
    # ============================================
    functions_by_id = {
        f.id: {
            "name": f.name,
            "number": f.function_number
        }
        for f in DependencyFunction.query.all()
    }

    # ============================================
    # SORT BY FUNCTION NUMBER  ✅ FIX
    # ============================================
    for dep_id, dep_studies in studies_by_dep.items():
        dep_studies.sort(
            key=lambda s: (
                int(functions_by_id[s.funciones_especificas]["number"])
                if s.funciones_especificas in functions_by_id
                and str(functions_by_id[s.funciones_especificas]["number"]).isdigit()
                else 9999
            )
        )

    # ============================================
    # SERIES MAPS
    # ============================================
    series_by_id = {s.id: s.name for s in Series.query.all()}
    subseries_by_id = {s.id: s.name for s in Subseries.query.all()}
    doctypes_by_id = {d.id: d.name for d in Doctype.query.all()}

    # ============================================
    # RENDER
    # ============================================
    return render_template(
        'list_documental_studies.html',
        studies=studies,
        dependencies_hierarchy=dependencies_hierarchy,
        studies_by_dep=studies_by_dep,
        functions_by_id=functions_by_id,
        series_by_id=series_by_id,
        subseries_by_id=subseries_by_id,
        doctypes_by_id=doctypes_by_id,
        entities=entities,
        active_entity_id=active_entity_id
    )

@app.route("/api/create_doctype_chain", methods=["POST"])
def api_create_doctype_chain():
    try:
        data = request.get_json(silent=True) or {}

        print("DATA RECEIVED:", data)

        # ---------------------------
        # INPUTS
        # ---------------------------
        serie_name = (data.get("serie") or "").strip()
        subserie_name = (data.get("subserie") or "").strip()
        doctype_name = (data.get("doctype") or "").strip()

        if not serie_name or not subserie_name or not doctype_name:
            return jsonify(success=False, error="Campos obligatorios"), 400

        # Normalize for comparison
        serie_norm = serie_name.lower()
        subserie_norm = subserie_name.lower()
        doctype_norm = doctype_name.lower()

        # ===========================
        # SERIE
        # ===========================
        serie = Series.query.filter(
            func.lower(Series.name) == serie_norm
        ).first()

        if not serie:
            serie = Series(
                name=serie_name,
                definition="N/A"  # 🔥 REQUIRED FIX
            )
            db.session.add(serie)
            db.session.flush()

        # ===========================
        # SUBSERIE
        # ===========================
        subserie = Subseries.query.filter(
            func.lower(Subseries.name) == subserie_norm,
            Subseries.series_id == serie.id
        ).first()

        if not subserie:
            subserie = Subseries(
                name=subserie_name,
                series_id=serie.id,
                definition = data.get("definition") or "N/A"
            )
            db.session.add(subserie)
            db.session.flush()

        # ===========================
        # DOCTYPE
        # ===========================
        doctype = Doctype.query.filter(
            func.lower(Doctype.name) == doctype_norm,
            Doctype.subseries_id == subserie.id
        ).first()

        if not doctype:
            doctype = Doctype(
                name=doctype_name,
                subseries_id=subserie.id,
                definition="N/A"   # 🔥 REQUIRED FIX
            )
            db.session.add(doctype)
            db.session.flush()

        # ===========================
        # COMMIT
        # ===========================
        db.session.commit()

        return jsonify(
            success=True,
            serie_id=serie.id,
            subserie_id=subserie.id,
            doctype_id=doctype.id
        )

    except Exception as e:
        db.session.rollback()
        print("🔥 ERROR in /api/create_doctype_chain:", e)
        import traceback
        traceback.print_exc()

        return jsonify(success=False, error=str(e)), 500

# ========================================================
# GET USED FUNCTIONS FOR dependencies (AJAX)
# ========================================================
@app.get("/api/dependencies/<int:dependency_id>/used-functions")
def get_used_functions(dependency_id):
    used = (
        db.session.query(DocumentalStudy.funciones_especificas)
        .filter(DocumentalStudy.dependency_id == dependency_id)
        .filter(DocumentalStudy.funciones_especificas.isnot(None))
        .all()
    )

    used_ids = [u[0] for u in used]

    return jsonify(used_ids)

# =========================================================
# ADD DOCUMENTAL STUDY
# =========================================================
@app.route('/add_documental_studies', methods=['GET', 'POST'])
def add_documental_studies():

    entities = Entity.query.order_by(Entity.name).all()

    if request.method == 'POST':
        try:
            # ---------- Required ----------
            entity_id = request.form.get("entity_id", type=int)
            dependency_id = request.form.get("dependency_id", type=int)

            func_raw = request.form.get("funciones_especificas")
            func_id = int(func_raw) if func_raw and func_raw.isdigit() else None

            serie_raw = request.form.get("serie_id")
            subserie_raw = request.form.get("subserie_id")
            doctype_raw = request.form.get("tipo_documental_id")

            serie_id = int(serie_raw) if serie_raw and serie_raw.isdigit() else None
            subserie_id = int(subserie_raw) if subserie_raw and subserie_raw.isdigit() else None
            doctype_id = int(doctype_raw) if doctype_raw and doctype_raw.isdigit() else None

            nombre_expediente = request.form.get("nombre_expediente")

            # ---------- VALIDATION ----------
            if not entity_id:
                return _handle_response_error("Debe seleccionar una Entidad.")

            if not dependency_id:
                return _handle_response_error("Debe seleccionar una Dependencia.")

            if not func_id:
                return _handle_response_error("Debe seleccionar una Función específica.")

            if not serie_raw or not subserie_raw or not doctype_raw:
                return _handle_response_error(
                    "Debe seleccionar Serie, Subserie y Tipo Documental."
                )

            if not nombre_expediente:
                return _handle_response_error(
                    "Debe ingresar el nombre del expediente."
                )

            # ===========================
            # SAFE FIELD MAPPING
            # ===========================
            allowed_fields = {c.name for c in DocumentalStudy.__table__.columns}
            data = {}
            # ===========================
            # GENERIC FIELD MAPPING
            # ===========================
            for field in allowed_fields:

                if field == "id":
                    continue

                if field in [
                    "entity_id",
                    "dependency_id",
                    "serie_id",
                    "subserie_id",
                    "tipo_documental_id",
                    "funciones_especificas",
                    "non_documented"
                ]:
                    continue

                data[field] = normalize_na(
                    request.form.get(field)
                )
            # ===========================
            # FORCE IDS
            # ===========================
            data["entity_id"] = entity_id
            data["dependency_id"] = dependency_id
            data["serie_id"] = serie_id
            data["subserie_id"] = subserie_id
            data["tipo_documental_id"] = doctype_id
            data["funciones_especificas"] = func_id
            data["non_documented"] = bool(request.form.get("non_documented"))
            # =====================================
            # SOPORTE / FORMATO
            # =====================================
            data["spt_fisico"] = request.form.get("spt_fisico")
            data["spt_digital"] = request.form.get("spt_digital")

            data["spt_url"] = request.form.get("spt_url")
            data["spt_otro"] = request.form.get("spt_otro")

            data["spt_formInicial"] = request.form.get("spt_formInicial")
            data["spt_formFinal"] = request.form.get("spt_formFinal")

            # ===========================
            # SAVE MAIN STATE
            # ===========================
            session["form_main"] = {
                "entity_id": request.form.get("entity_id"),
                "dependency_id": request.form.get("dependency_id"),
                "acto_administrativo": request.form.get("acto_administrativo"),
                "nombre_dependencia": request.form.get("nombre_dependencia"),
                "nombre_grupo": request.form.get("nombre_grupo"),
                "lider_dependencia": request.form.get("lider_dependencia"),
                "funcionario_entrevistado": request.form.get("funcionario_entrevistado"),
            }

            # ===========================
            # 🔥 SAVE AUX STATE (FIX)
            # ===========================
            main_keys = set(session["form_main"].keys())

            session["form_aux"] = {
                key: request.form.get(key)
                for key in request.form
                if key not in main_keys
            }

            # ===========================
            # 🔥 SAVE SELECTED IDS (FIX)
            # ===========================
            session["selected_serie"] = serie_id
            session["selected_subserie"] = subserie_id
            session["selected_doctype"] = doctype_id
            session["selected_funcion"] = func_id

            # ===========================
            # SAVE LAST FUNCIONARIO
            # ===========================
            funcionario = request.form.get("funcionario_entrevistado")
            if funcionario:
                session["last_funcionario_entrevistado"] = funcionario

            # ===========================
            # ND LOGIC
            # ===========================
            if data["non_documented"]:
                serie = get_or_create_na(Series)
                subserie = get_or_create_na(Subseries, series_id=serie.id)
                doctype = get_or_create_na(Doctype, subseries_id=subserie.id)

                data["serie_id"] = serie.id
                data["subserie_id"] = subserie.id
                data["tipo_documental_id"] = doctype.id

                if not data.get("nombre_expediente"):
                    data["nombre_expediente"] = "N/A"

                if not data.get("comentarios_entrevista"):
                    data["comentarios_entrevista"] = "N/A"

            # ===========================
            # CREATE
            # ===========================
            study = DocumentalStudy(**data)
            study.nombre_expediente = study.nombre_expediente or "N/A"
            study.comentarios_entrevista = study.comentarios_entrevista or "N/A"

            db.session.add(study)

            # ===========================
            # UPDATE FUNCTION FLAG
            # ===========================
            from models import DependencyFunction

            if func_id:
                df = DependencyFunction.query.get(func_id)
                if df:
                    df.is_non_documented = True

            db.session.commit()

            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify(success=True)

            flash("✅ Estudio Documental guardado correctamente.", "success")

            return redirect(url_for(
                'add_documental_studies',
                entity_id=entity_id,
                dependency_id=dependency_id
            ))

        except Exception as e:
            db.session.rollback()
            return _handle_response_error(f"Error al guardar: {e}")

    # ===========================
    # GET (CLEAN FIX)
    # ===========================
    saved_dependency_id = request.args.get("dependency_id", type=int)
    entity_id = request.args.get("entity_id", type=int)

    study = None

    # 🔥 LOAD LAST CREATED STUDY (same context)
    if entity_id and saved_dependency_id:
        study = (
            DocumentalStudy.query
            .filter_by(entity_id=entity_id, dependency_id=saved_dependency_id)
            .order_by(DocumentalStudy.id.desc())
            .first()
        )

    form_main = {
        "entity_id": entity_id,
        "dependency_id": saved_dependency_id,
        "acto_administrativo": (
            study.acto_administrativo if study else None
        ),
        "nombre_dependencia": (
            study.nombre_dependencia if study else None
        ),
        "nombre_grupo": (
            study.nombre_grupo if study else None
        ),
        "lider_dependencia": (
            study.lider_dependencia if study else None
        ),
    }

    form_aux = study.__dict__ if study else {}

    return render_template(
        "add_documental_studies.html",
        entities=entities,
        form_main=form_main,
        form_aux=form_aux,
        last_funcionario=session.get("last_funcionario_entrevistado", ""),
        savedDependencyId=saved_dependency_id,

        selected_serie=study.serie_id if study else None,
        selected_subserie=study.subserie_id if study else None,
        selected_doctype=study.tipo_documental_id if study else None,
        selected_funcion=study.funciones_especificas if study else None,
    )

def _handle_response_error(msg, field_errors=None):
    """
    Consistent error response:
    - JSON (AJAX): message + field_errors
    - Non-AJAX: flash only
    """
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        payload = {
            "success": False,
            "message": msg,
        }
        if field_errors:
            payload["field_errors"] = field_errors
        return jsonify(payload), 400

    flash(msg, "danger")
    return redirect(url_for("add_documental_studies"))


@app.route('/edit_documental_studies/<int:id>', methods=['GET', 'POST'])
def edit_documental_studies(id):
    study = DocumentalStudy.query.get_or_404(id)

    if request.method == 'POST':
        try:
            funciones_value = request.form.get('funciones_especificas')
            funciones_id = int(funciones_value) if funciones_value else None
            study.version = request.form.get('version')
            study.entity_id = request.form.get('entity_id', type=int)
            study.dependency_id = request.form.get('dependency_id', type=int)
            study.acto_administrativo = request.form.get('acto_administrativo')
            study.nombre_dependencia = request.form.get('nombre_dependencia')
            study.nombre_grupo = request.form.get('nombre_grupo')
            study.formalizacion_entrevista = request.form.get('formalizacion_entrevista')
            study.fecha_entrevista = request.form.get('fecha_entrevista')
            study.lider_dependencia = request.form.get('lider_dependencia')
            study.funcionario_entrevistado = request.form.get('funcionario_entrevistado')
            study.funciones_especificas = funciones_id
            study.non_documented = bool(request.form.get('non_documented'))
            study.comentarios_entrevista = request.form.get('comentarios_entrevista')
            study.nombre_expediente = request.form.get('nombre_expediente')
            study.serie_id = request.form.get('serie_id', type=int)
            study.subserie_id = request.form.get('subserie_id', type=int)
            study.tipo_documental_id = request.form.get('tipo_documental_id', type=int)
            study.spt_fisico = request.form.get('spt_fisico')
            study.spt_digital = request.form.get('spt_digital')
            study.spt_url = request.form.get('spt_url')
            study.spt_otro = request.form.get('spt_otro')
            study.spt_formInicial = request.form.get('spt_formInicial')
            study.spt_formFinal = request.form.get('spt_formFinal')
            study.periodicidad_expediente = request.form.get('periodicidad_expediente')
            study.volumen_documental_anno = request.form.get('volumen_documental_anno')
            study.estado_organizacion_archivos = request.form.get('estado_organizacion_archivos')
            study.nombre_software = request.form.get('nombre_software')
            study.dependencia_grupo_tramite = request.form.get('dependencia_grupo_tramite')
            study.dependencia_grupo_consulta = request.form.get('dependencia_grupo_consulta')
            study.norma_interna = request.form.get('norma_interna')
            study.norma_externa = request.form.get('norma_externa')
            study.tramite_original = request.form.get('tramite_original')
            study.tramite_copia = request.form.get('tramite_copia')
            study.tiempo_conservacion_expediente = request.form.get('tiempo_conservacion_expediente')
            study.trd_annosOficina = request.form.get('trd_annosOficina')
            study.trd_annosTransfAC = request.form.get('trd_annosTransfAC')
            study.trd_perTransferencias = request.form.get('trd_perTransferencias')
            study.infoCaracter_publico = request.form.get('infoCaracter_publico')
            study.infoCaracter_privado = request.form.get('infoCaracter_privado')
            study.infoCaracter_confidencial = request.form.get('infoCaracter_confidencial')
            study.infoCaracter_usoInterno = request.form.get('infoCaracter_usoInterno')
            study.orc_delegadoActAdmitivo = request.form.get('orc_delegadoActAdmitivo')
            study.orc_participante = request.form.get('orc_participante')
            study.sgc_nombre_proceso = request.form.get('sgc_nombre_proceso')
            study.sgc_nombre_procedimiento = request.form.get('sgc_nombre_procedimiento')
            study.prodDerechosHumanos = request.form.get('prodDerechosHumanos')
            study.observaciones = request.form.get('observaciones')
            if study.non_documented:
                if not study.nombre_expediente:
                    study.nombre_expediente = "N/A"

                if not study.comentarios_entrevista:
                    study.comentarios_entrevista = "N/A"
            db.session.commit()
            flash("Estudio Documental actualizado correctamente.", "success")
            return redirect(url_for('list_documental_studies'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error al actualizar: {e}", "danger")

    entities = Entity.query.order_by(Entity.name).all()
    dependencies = Dependency.query.order_by(Dependency.name).all()
    series = Series.query.order_by(Series.name).all()
    subseries = Subseries.query.order_by(Subseries.name).all()
    doctypes = Doctype.query.order_by(Doctype.name).all()

    #return render_template(
    #    'edit_documental_studies.html',
    #    study=study,
    #    entities=entities,
    #    dependencies=dependencies,
    #    series=series,
    #    subseries=subseries,
    #    doctypes=doctypes,
    #)

    entities = Entity.query.order_by(Entity.name).all()

    return render_template(
        "edit_documental_studies.html",
        study=study,
        entities=entities,
    )

# =========================================================
# DELETE DOCUMENTAL STUDY (with function unmarking logic)
# =========================================================
@app.route('/delete_documental_studies/<int:id>')
def delete_documental_studies(id):

    study = DocumentalStudy.query.get_or_404(id)

    function_id = study.funciones_especificas

    # FIRST delete the study
    db.session.delete(study)
    db.session.flush()

    # CHECK if another study still uses the same function
    if function_id:

        still_used = DocumentalStudy.query.filter(
            DocumentalStudy.funciones_especificas == function_id
        ).first()

        # If no more studies use it → unmark
        if not still_used:

            dependency_function = DependencyFunction.query.get(function_id)

            if dependency_function:
                dependency_function.is_non_documented = False

    try:
        db.session.commit()
        flash('Estudio documental eliminado correctamente.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('list_documental_studies'))
# =========================================================
# TRD EXCEL FORMATTER
# =========================================================
def apply_trd_format(ws, max_col=8, start_data_row=2):
    """
    Applies professional TRD formatting to worksheet.
    Safe for generator_code_excel.
    """

    bold = Font(bold=True, name="Calibri", size=11)
    title_font = Font(bold=True, name="Calibri", size=14)
    normal = Font(bold=False, name="Calibri", size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(border_style="thin", color="000000")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)

    fill_header = PatternFill("solid", fgColor="D9D9D9")

    # ==========================================
    # HEADER ROW
    # ==========================================
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.fill = fill_header
        cell.alignment = center
        cell.border = border_all

    # ==========================================
    # DATA ROWS
    # ==========================================
    last_row = ws.max_row

    from openpyxl.cell.cell import MergedCell

    for r in range(start_data_row, last_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)

            # 🚫 Skip merged cells
            if isinstance(cell, MergedCell):
                continue

            cell.font = normal
            cell.border = border_all

            if c in (1, 4, 5, 6):
                cell.alignment = left
            else:
                cell.alignment = center

            # Text columns left aligned
            if c in (1, 4, 5, 6):
                cell.alignment = left
            else:
                cell.alignment = center

    # ==========================================
    # COLUMN WIDTHS (balanced)
    # ==========================================
    widths = {
        1: 20,  # Código
        2: 10,  # Serie index
        3: 10,  # Subserie index
        4: 30,  # Serie name
        5: 30,  # Subserie name
        6: 35,  # Tipo documental
        7: 8,   # AG
        8: 8,   # AC
    }
    from openpyxl.utils import get_column_letter

    for col_idx, width in widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

# =========================================================
# GENERATOR HELPERS
# =========================================================

def _upper(s):
    return (s or "").strip().upper()


# =========================================================
# AJAX LEGACY HELPERS
# =========================================================
@app.route("/get_subseries/<int:serie_id>")
def get_subseries(serie_id):
    subseries = Subseries.query.filter_by(series_id=serie_id).all()

    return jsonify([
        {"id": s.id, "name": s.name}
        for s in subseries
    ])


@app.route("/get_doctypes/<int:subserie_id>")
def get_doctypes(subserie_id):
    try:
        doctypes = Doctype.query.filter_by(subseries_id=subserie_id).all()
    except Exception:
        doctypes = []
    return jsonify([{"id": d.id, "name": d.name} for d in doctypes])


@app.route('/get_dependencies_parent/<int:dependency_id>')
def get_dependencies_parent(dependency_id):
    dep = db.session.get(Dependency, dependency_id)
    if not dep or not dep.parent_id:
        return jsonify({'parent_name': None})
    parent = Dependency.query.get(dep.parent_id)
    return jsonify({'parent_name': parent.name if parent else None})


@app.get('/get_dependencies_functions/<int:dependency_id>')
def get_dependencies_functions(dependency_id):
    functions = DependencyFunction.query.filter_by(dependency_id=dependency_id).all()

    return jsonify([
        {
            "id": f.id,
            "name": f.name,
            "function_number": f.function_number,
            "non_documented": bool(f.is_non_documented),
        }
        for f in functions
    ])


@app.route('/get_function_non_documented/<int:function_id>')
def get_function_non_documented(function_id):
    func = DependencyFunction.query.get(function_id)
    if func:
        return jsonify({'non_documented': func.is_non_documented or 0})
    return jsonify({'non_documented': 0})


@app.route('/toggle_function_non_documented/<int:function_id>', methods=['POST'])
def toggle_function_non_documented(function_id):
    """Toggle DependencyFunction.non_documented (0 ↔ 1)."""
    func = DependencyFunction.query.get(function_id)
    if not func:
        return jsonify({'success': False, 'error': 'Function not found'}), 404
    func.is_non_documented = 0 if func.is_non_documented else 1
    db.session.commit()
    return jsonify({
        'success': True,
        'non_documented': func.is_non_documented,
    })

# =========================================================
# dependencies FUNCTIONS API
# =========================================================
@app.get("/api/functions")
def api_functions():

    dependency_id = request.args.get("dependency_id", type=int)
    entity_id = request.args.get("entity_id", type=int)

    # ===============================
    # CASE 1: centralized entity
    # ===============================
    if entity_id:
        entity = Entity.query.get(entity_id)

        if entity and entity.is_centralized and entity.central_dependency_code:
            dep = Dependency.query.filter_by(
                entity_id=entity.id,
                code=entity.central_dependency_code
            ).first()

            if dep:
                dependency_id = dep.id

    # ===============================
    # FINAL VALIDATION
    # ===============================
    if not dependency_id:
        return jsonify([])

    # ===============================
    # QUERY FUNCTIONS
    # ===============================
    functions = (
        DependencyFunction.query
        .filter_by(dependency_id=dependency_id)
        .order_by(DependencyFunction.function_number.asc())
        .all()
    )

    return jsonify([
        {
            "id": f.id,
            "name": f.name,
            "function_number": f.function_number,
            "is_non_documented": f.is_non_documented,
            "is_used": False  # optional, keeps frontend stable
        }
        for f in functions
    ])    
# =========================================================
# BULK INSERT BLUEPRINT
# =========================================================
bulk_bp = Blueprint('bulk_import', __name__)


@bulk_bp.route("/bulk_insert", methods=["GET", "POST"])
def bulk_insert():
    from models import MODEL_MAP
    if request.method == "POST":
        model_name = request.form.get("model")
        column_name = request.form.get("column")
        raw_text = request.form.get("data")
        if not model_name or not column_name or not raw_text:
            flash("Todos los campos son obligatorios", "danger")
            return redirect(url_for("bulk_import.bulk_insert"))

        model = MODEL_MAP.get(model_name)
        if not model:
            flash("Modelo inválido", "danger")
            return redirect(url_for("bulk_import.bulk_insert"))

        lines = [ln.strip() for ln in raw_text.splitlines() if ln.strip()]
        for ln in lines:
            obj = model(**{column_name: ln})
            db.session.add(obj)
        db.session.commit()
        flash(f"{len(lines)} records added.", "success")
        return redirect(url_for("bulk_import.bulk_insert"))

    entities = Entity.query.all()
    dependencies = Dependency.query.all()
    return render_template("bulk_insert.html", entities=entities, dependencies=dependencies)


app.register_blueprint(bulk_bp)

# =========================================================
# TRD EXCEL GENERATOR
# =========================================================

# =========================================================
# EXCEL EXPORT — BREAK BY dependencies
# =========================================================
@app.route('/generator_code_excel')
def generator_code_excel():

    from collections import defaultdict
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    entity_id = request.args.get("entity_id", type=int)

    if not entity_id:
        abort(400, "Entity required")

    entity = Entity.query.get_or_404(entity_id)

    # ---------------------------
    # GET STUDIES
    # ---------------------------
    studies = (
        DocumentalStudy.query
        .filter(DocumentalStudy.entity_id == entity_id)
        .options(
            joinedload(DocumentalStudy.dependency),
            joinedload(DocumentalStudy.serie),
            joinedload(DocumentalStudy.subserie),
            joinedload(DocumentalStudy.doctype)
        )
        .all()
    )

    # ---------------------------
    # GROUP BY dependencies
    # ---------------------------
    studies_by_dep = defaultdict(list)

    for s in studies:
        if s.dependency_id:
            studies_by_dep[s.dependency_id].append(s)

    # ---------------------------
    # FUNCTIONS MAP
    # ---------------------------
    functions_by_id = {
        f.id: f.function_number
        for f in DependencyFunction.query.all()
    }

    # ---------------------------
    # FILTER + SORT (same as HTML)
    # ---------------------------
    for dep_id in list(studies_by_dep.keys()):

        dep_studies = [
            s for s in studies_by_dep[dep_id]
            if not getattr(s, "non_documented", False)
        ]

        dep_studies.sort(
            key=lambda s: (
                int(functions_by_id[s.funciones_especificas])
                if s.funciones_especificas in functions_by_id
                and str(functions_by_id[s.funciones_especificas]).isdigit()
                else 9999
            )
        )

        studies_by_dep[dep_id] = dep_studies

    # ---------------------------
    # WORKBOOK
    # ---------------------------
    wb = Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for dep_id in sorted(studies_by_dep.keys()):

        dep_studies = studies_by_dep[dep_id]

        if not dep_studies:
            continue

        dep = Dependency.query.get(dep_id)
        if not dep:
            continue

        # 🔥 uses fixed builder
        estructura = build_tvd_structure(dep_studies, entity)

        ws = wb.create_sheet(title=dep.name[:31])

        row = 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        ws.cell(row=row, column=1, value=f"ENTIDAD: {entity.name}").font = bold
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        ws.cell(row=row, column=1, value=f"DEPENDENCIA: {dep.name}").font = bold
        row += 2

        headers = [
            "CÓDIGO", "SERIE / SUBSERIE / DOCUMENTO",
            "P", "E", "RETENCIÓN",
            "C", "S", "E",
            "REPRODUCCIÓN", "DDHH/DIH",
            "FUNCIÓN", "PROCEDIMIENTO"
        ]

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.alignment = center

        row += 1

        # ---------------------------
        # DATA
        # ---------------------------
        for serie in estructura:

            ws.cell(row=row, column=1, value=serie["codigo"])
            ws.cell(row=row, column=2, value=serie["nombre"])

            row += 1

            for sub in serie["subseries"]:

                ws.cell(row=row, column=1, value=sub["codigo"])
                ws.cell(row=row, column=2, value="   " + sub["nombre"])

                row += 1

                for doc in sub["doctypes"]:

                    ws.cell(row=row, column=1, value=doc["codigo"])
                    ws.cell(row=row, column=2, value="      " + doc["nombre"])

                    ws.cell(row=row, column=3, value="X" if doc["spt_fisico"] else "")
                    ws.cell(row=row, column=4, value="X" if doc["spt_digital"] else "")
                    ws.cell(row=row, column=5, value=doc["trd_annosOficina"] or "") 
                    ws.cell(row=row, column=6, value=doc["trd_annosTransfAC"] or "")    
                    ws.cell(row=row, column=11,
                            value=doc["function_number"] + " - " + doc["function_name"]
                            if doc["function_number"] else "")

                    ws.cell(row=row, column=12, value=doc["procedimiento"])

                    row += 1

        widths = [15, 45, 6, 6, 12, 5, 5, 5, 12, 10, 35, 45]

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"TVD_{entity.name}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================================================
# DATA TRANSFORMATION   (FLAT → HIERARCHY)
# =========================================================
def build_tvd_structure_from_flat(data):

    estructura = {}

    for d in data:

        serie_name = d["serie"] or "N/A"
        subserie_name = d["subserie"] or "N/A"

        if serie_name not in estructura:
            estructura[serie_name] = {
                "nombre": serie_name,
                "codigo": "",
                "disposicion": "",
                "subseries": {}
            }

        serie = estructura[serie_name]

        if subserie_name not in serie["subseries"]:
            serie["subseries"][subserie_name] = {
                "nombre": subserie_name,
                "codigo": "",
                "disposicion": "",
                "doctypes": []
            }

        sub = serie["subseries"][subserie_name]

        # ✅ SAFE EXTRACTION FROM FLAT DATA
        function_number = d.get("function_number", "")
        function_name = d.get("function_name", "")

        sub["doctypes"].append({
            "codigo": d["codigo"],
            "nombre": d["doctype"],

            "spt_fisico": d["spt_fisico"],
            "spt_digital": d["spt_digital"],

            # ✅ RETENTION
            "trd_annosOficina": d.get("trd_annosOficina", ""),
            "trd_annosTransfAC": d.get("trd_annosTransfAC", ""),

            "disposicion": "",

            "reproduccion_tecnica": d["reproduccion_tecnica"],
            "es_ddhh": d["es_ddhh"],

            # ✅ FIXED (NOW DEFINED)
            "function_number": function_number,
            "function_name": function_name,

            "procedimiento": d["procedimiento"]
        })

    estructura_list = []

    for serie in estructura.values():
        serie["subseries"] = list(serie["subseries"].values())
        estructura_list.append(serie)

    return estructura_list

# =========================================================
# DATA RETRIEVAL WITH FUNCTION JOIN (REQUIRED)
# =========================================================

from sqlalchemy import func
from sqlalchemy.orm import aliased

def get_tvd_data(entity_id):

    entity = Entity.query.get(entity_id)

    if not entity:
        return []

    separator = entity.separator or "."

    # ==========================================
    # 🔥 CENTRALIZED SQL LOGIC
    # ==========================================
    if entity.is_centralized and entity.central_dependency_code:
        dep_code_expr = func.trim(entity.central_dependency_code)
    else:
        dep_code_expr = func.trim(Dependency.code)

    # ==========================================
    # 🔥 BASE QUERY (FIXED WITH FUNCTION JOIN)
    # ==========================================
    studies = (
    db.session.query(
        DocumentalStudy.id,

        dep_code_expr.label("dep_code"),

        Series.name.label("serie_name"),
        Subseries.name.label("subserie_name"),
        Doctype.name.label("doctype_name"),

        DocumentalStudy.spt_fisico,
        DocumentalStudy.spt_digital,
        DocumentalStudy.trd_annosTransfAC,
        DocumentalStudy.reproduccion_tecnica,
        DocumentalStudy.es_ddhh,
        DocumentalStudy.procedimiento,

        DependencyFunction.function_number.label("function_number"),
        DependencyFunction.name.label("function_name"),
    )
    .join(Dependency, DocumentalStudy.dependency_id == Dependency.id)
    .join(Series, DocumentalStudy.serie_id == Series.id)
    .outerjoin(Subseries, DocumentalStudy.subserie_id == Subseries.id)
    .outerjoin(Doctype, DocumentalStudy.tipo_documental_id == Doctype.id)
    .outerjoin(
        DependencyFunction,
        DocumentalStudy.funciones_especificas == DependencyFunction.id
    )

    # ✅ FILTER ADDED HERE
    .filter(
        DocumentalStudy.entity_id == entity_id,
        Series.name != "N/A"
    )

    .all()
)
    # ==========================================
    # INDEXES
    # ==========================================
    serie_names = sorted({s.serie_name for s in studies if s.serie_name})
    subserie_names = sorted({s.subserie_name for s in studies if s.subserie_name})

    serie_index_map = {
        name: str(i + 1).zfill(2)
        for i, name in enumerate(serie_names)
    }

    subserie_index_map = {
        name: str(i + 1).zfill(2)
        for i, name in enumerate(subserie_names)
    }

    # ==========================================
    # FINAL RESULT
    # ==========================================
    result = []

    for s in studies:

        serie_idx = serie_index_map.get(s.serie_name, "00")
        sub_idx = subserie_index_map.get(s.subserie_name, "00")

        code = f"{s.dep_code}{separator}{serie_idx}{separator}{sub_idx}"

        result.append({
            "codigo": code,
            "serie": s.serie_name,
            "subserie": s.subserie_name,
            "doctype": s.doctype_name,

            "spt_fisico": s.spt_fisico,
            "spt_digital": s.spt_digital,

            # ✅ REQUIRED
            "trd_annosOficina": s.trd_annosOficina,
            "trd_annosTransfAC": s.trd_annosTransfAC,
            "reproduccion_tecnica": s.reproduccion_tecnica,
            "es_ddhh": s.es_ddhh,
            "procedimiento": s.procedimiento,

            # 🔥 FIXED OUTPUT
            "function_number": s.function_number or "",
            "function_name": s.function_name or "",
        })

    return result

# =====================================================
# GENERATOR CODE (HTML VIEW)
# =====================================================
# =========================================================
# TVD GENERATOR (BREAK BY dependencies)
# =========================================================
@app.route('/generator_code')
@app.route('/generator_code')
def generator_code():

    from collections import defaultdict
    from sqlalchemy.orm import joinedload

    entity_id = request.args.get("entity_id", type=int)

    entities = Entity.query.order_by(Entity.name).all()

    # ---------------------------
    # NO ENTITY SELECTED
    # ---------------------------
    if not entity_id:
        return render_template(
            "code_generator.html",
            entities=entities,
            estructura_by_dep=None,
            selected_entity_id=None
        )

    entity = Entity.query.get(entity_id)

    if not entity:
        abort(404)

    # ---------------------------
    # GET STUDIES (OPTIMIZED)
    # ---------------------------
    studies = (
        DocumentalStudy.query
        .filter(DocumentalStudy.entity_id == entity_id)
        .options(
            joinedload(DocumentalStudy.serie),
            joinedload(DocumentalStudy.subserie),
            joinedload(DocumentalStudy.doctype)
        )
        .all()
    )

    # ---------------------------
    # GROUP BY dependencies
    # ---------------------------
    studies_by_dep = defaultdict(list)

    for s in studies:
        if s.dependency_id:
            studies_by_dep[s.dependency_id].append(s)

    # ---------------------------
    # FUNCTIONS MAP (FOR SORT)
    # ---------------------------
    functions_by_id = {
        f.id: f.function_number
        for f in DependencyFunction.query.all()
    }

    # ---------------------------
    # SORT + FILTER NON-DOCUMENTED ✅
    # ---------------------------
    for dep_id in list(studies_by_dep.keys()):

        dep_studies = studies_by_dep[dep_id]

        # 🔴 REMOVE non_documented from structure/code generation
        dep_studies = [
            s for s in dep_studies
            if not getattr(s, "non_documented", False)
        ]

        # SORT by function_number
        dep_studies.sort(
            key=lambda s: (
                int(functions_by_id[s.funciones_especificas])
                if s.funciones_especificas in functions_by_id
                and str(functions_by_id[s.funciones_especificas]).isdigit()
                else 9999
            )
        )

        # overwrite with cleaned + sorted list
        studies_by_dep[dep_id] = dep_studies

    # ---------------------------
    # BUILD STRUCTURE PER DEP
    # ---------------------------
    estructura_by_dep = {}

    for dep_id in sorted(studies_by_dep.keys()):
        dep_studies = studies_by_dep[dep_id]

        if not dep_studies:
            continue  # skip empty after filtering

        dep = Dependency.query.get(dep_id)

        if not dep:
            continue

        estructura = build_tvd_structure(dep_studies, entity)

        # KEY = dependency object (used in template)
        estructura_by_dep[dep] = estructura

    # ---------------------------
    # SPECIAL DEPENDENCIES
    # ---------------------------
    tvd_dependency = Dependency.query.filter_by(
        entity_id=entity_id,
        tvd_responsible=True
    ).first()

    legal_dependency = Dependency.query.filter_by(
        entity_id=entity_id,
        legal_representante=True
    ).first()

    # ---------------------------
    # RENDER
    # ---------------------------
    return render_template(
        "code_generator.html",
        entities=entities,
        estructura_by_dep=estructura_by_dep,
        selected_entity_id=entity_id,
        entidad_nombre=entity.name,
        tvd_dependency=tvd_dependency,
        legal_dependency=legal_dependency,
        entity=entity,
        get_logo_url=get_logo_url
    )

# =========================================================
# SHARED LOGIC FOR HTML + EXCEL (ONLY SOURCE)
# =========================================================
def _get_generator_payload(selected_entity_id):

    estructura = None
    entidad_nombre = None
    dependency = None
    tvd_dependency = None
    legal_dependency = None

    if selected_entity_id:

        entidad = db.session.get(Entity, selected_entity_id)

        if not entidad:
            raise ValueError("Entidad no encontrada")

        entidad_nombre = entidad.name

        # 🔥 SAME SOURCE AS HTML
        data = get_tvd_data(selected_entity_id)
        estructura = build_tvd_structure_from_flat(data)

        tvd_dependency = (
            db.session.query(Dependency)
            .filter(
                Dependency.entity_id == selected_entity_id,
                Dependency.tvd_responsible == True
            )
            .order_by(Dependency.id)
            .first()
        )

        legal_dependency = (
            db.session.query(Dependency)
            .filter(
                Dependency.entity_id == selected_entity_id,
                Dependency.legal_representante == True
            )
            .order_by(Dependency.id)
            .first()
        )

        if entidad.is_centralized:
            dependency = (
                db.session.query(Dependency)
                .filter_by(
                    entity_id=entidad.id,
                    code=entidad.central_dependency_code
                )
                .first()
            )

        return {
            "estructura": estructura,
            "entidad_nombre": entidad_nombre,
            "dependency": dependency,
            "tvd_dependency": tvd_dependency,
            "legal_dependency": legal_dependency,
            "entity": entidad
        }

    return None


# =========================================================
# ENTITY DOCTYPES VIEW (AJAX + LEGACY)
# =========================================================
@app.route("/entity_doctypes")
def entity_doctypes():

    entity_id = request.args.get("entity_id", type=int) \
        or session.get("active_entity_id")

    if not entity_id:
        return redirect(url_for("list_entities"))

    entity = Entity.query.get_or_404(entity_id)

    # Ensure ordering tables exist
    db.session.commit()

    studies = (
        DocumentalStudy.query
        .filter_by(entity_id=entity_id)
        .all()
    )

    # --------------------------------------
    # MAPS
    # --------------------------------------
    series_by_id = {s.id: s.name for s in Series.query.all()}
    subseries_by_id = {s.id: s.name for s in Subseries.query.all()}
    doctypes_by_id = {d.id: d.name for d in Doctype.query.all()}
    dependencies = {d.id: d for d in Dependency.query.all()}

# --------------------------------------
#   BUILD HIERARCHICAL STRUCTURE
# --------------------------------------
def build_tvd_structure(studies, entity):

    from models import DependencyFunction
    import re
    import unicodedata

    # =====================================
    # HELPERS
    # =====================================
    def normalize_name(text):
        text = (text or "").strip().lower()

        text = unicodedata.normalize("NFKD", text)
        text = "".join(
            c for c in text
            if not unicodedata.combining(c)
        )

        text = re.sub(r"[^\w\s]", "", text)
        text = re.sub(r"\s+", " ", text).strip()

        if text.endswith("es") and len(text) > 4:
            text = text[:-2]
        elif text.endswith("s") and len(text) > 3:
            text = text[:-1]

        return text

    def normalize(text):
        return (text or "").strip().lower()

    functions_map = {
        f.id: f
        for f in DependencyFunction.query.all()
    }

    separator = entity.separator or "-"
    estructura = {}

    # =====================================
    # UNIQUE ORDERS
    # =====================================
    unique_series = []
    unique_series_norm = []

    unique_subseries = {}
    unique_subseries_norm = {}

    # =====================================
    # PREPASS
    # =====================================
    for s in studies:

        if not s.serie:
            continue

        serie_name = s.serie.name.strip()
        sub_name = s.subserie.name.strip() if s.subserie else "N/A"

        serie_norm = normalize_name(serie_name)
        sub_norm = normalize_name(sub_name)

        if serie_norm not in unique_series_norm:
            unique_series.append(serie_name)
            unique_series_norm.append(serie_norm)

        unique_subseries.setdefault(serie_norm, [])
        unique_subseries_norm.setdefault(serie_norm, [])

        if sub_norm not in unique_subseries_norm[serie_norm]:
            unique_subseries[serie_norm].append(sub_name)
            unique_subseries_norm[serie_norm].append(sub_norm)

    # =====================================
    # MAIN LOOP
    # =====================================
    for s in studies:

        if not s.serie:
            continue

        func = functions_map.get(s.funciones_especificas)

        function_number = func.function_number if func else ""
        function_name = func.name if func else ""

        serie_key = s.serie.name.strip()
        sub_key = s.subserie.name.strip() if s.subserie else "N/A"

        serie_norm = normalize_name(serie_key)
        sub_norm = normalize_name(sub_key)

        serie_order = unique_series_norm.index(serie_norm) + 1
        sub_order = unique_subseries_norm[serie_norm].index(sub_norm) + 1

        dep_code = (
            str(s.dependency.code).strip()
            if getattr(s, "dependency", None) and s.dependency.code
            else ""
        )

        serie_code = f"{dep_code}{separator}{serie_order}"
        sub_code = f"{serie_code}{separator}{sub_order}"
        doc_code = sub_code

        # ================= SERIE =================
        if serie_norm not in estructura:
            estructura[serie_norm] = {
                "codigo": serie_code,
                "nombre": serie_key,

                # documental_studies values
                "trd_annosOficina": s.trd_annosOficina,
                "trd_annosTransfAC": s.trd_annosTransfAC,
                "disposicion": s.disposicion_final,

                "reproduccion_tecnica": s.reproduccion_tecnica,
                "es_ddhh": s.prodDerechosHumanos,
                "procedimiento": getattr(s, "procedimiento", ""),

                "function_number": function_number,
                "function_name": function_name,

                "subseries": {}
            }

        # ================= SUBSERIE =================
        if sub_norm not in estructura[serie_norm]["subseries"]:
            estructura[serie_norm]["subseries"][sub_norm] = {
                "codigo": sub_code,
                "nombre": sub_key,

                # documental_studies values
                "trd_annosOficina": s.trd_annosOficina,
                "trd_annosTransfAC": s.trd_annosTransfAC,
                "disposicion": s.disposicion_final,

                "reproduccion_tecnica": s.reproduccion_tecnica,
                "es_ddhh": s.prodDerechosHumanos,
                "procedimiento": getattr(s, "procedimiento", ""),

                "function_number": function_number,
                "function_name": function_name,

                "doctypes": []
            }

        # ================= DOCTYPE =================
        doc_name = s.doctype.name if s.doctype else "N/A"
        doc_norm = normalize(doc_name)

        # 🔥 FIX: ALWAYS use normalized keys
        docs = estructura[serie_norm]["subseries"][sub_norm]["doctypes"]

        existing_docs = {
            normalize(d["nombre"])
            for d in docs
        }

        if doc_norm not in existing_docs:

            docs.append({
                "codigo": doc_code,
                "nombre": doc_name,

                # ✅ normalized booleans
                "spt_fisico": s.spt_fisico,
                "spt_digital": s.spt_digital,

                "trd_annosOficina": s.trd_annosOficina,
                "trd_annosTransfAC": s.trd_annosTransfAC,
                "disposicion": getattr(s, "disposicion_final", ""),
                "reproduccion_tecnica": s.reproduccion_tecnica,
                "es_ddhh": s.prodDerechosHumanos,
                "function_number": function_number,
                "function_name": function_name,
                "procedimiento": getattr(s, "procedimiento", "")
            })

        else:
            for d in docs:
                if normalize(d["nombre"]) == doc_norm:

                    if s.spt_fisico:
                        d["spt_fisico"] = s.spt_fisico

                    if s.spt_digital:
                        d["spt_digital"] = s.spt_digital

                    break

    # =====================================
    # FINAL LIST
    # =====================================
    result = []

    for serie in estructura.values():
        serie["subseries"] = list(serie["subseries"].values())
        result.append(serie)

    return result

# =========================================================
# HELPER GENERATOR_CODE (SQL QUERY + PYTHON LOGIC)
# =========================================================

@app.route("/generator_code_sql", methods=["GET"])
def generator_code_sql():

    entity_id = request.args.get("entity_id", type=int) \
        or session.get("active_entity_id")

    if not entity_id:
        return redirect(url_for("list_entities"))

    entity = Entity.query.get_or_404(entity_id)

    # ----------------------------------------------------
    # SQL QUERY (same logic you tested in MySQL Workbench)
    # ----------------------------------------------------
    sql = """
    SELECT

    ds.entity_id,
    d.code AS dependencies_code,
    df.function_number,

    (
        SELECT COUNT(DISTINCT LOWER(REPLACE(s2.name,' ','')))
        FROM documental_studies ds2
        JOIN series s2 ON s2.id = ds2.serie_id
        WHERE ds2.entity_id = ds.entity_id
          AND LOWER(REPLACE(s2.name,' ','')) 
              <= LOWER(REPLACE(s.name,' ',''))
    ) AS series_order,

    (
        SELECT COUNT(DISTINCT LOWER(REPLACE(ss2.name,' ','')))
        FROM documental_studies ds2
        JOIN subseries ss2 ON ss2.id = ds2.subserie_id
        WHERE ds2.entity_id = ds.entity_id
          AND LOWER(REPLACE(ss2.name,' ','')) 
              <= LOWER(REPLACE(ss.name,' ',''))
    ) AS subseries_order,

    (
        SELECT COUNT(DISTINCT LOWER(REPLACE(dt2.name,' ','')))
        FROM documental_studies ds2
        JOIN doctype dt2 ON dt2.id = ds2.tipo_documental_id
        WHERE ds2.entity_id = ds.entity_id
          AND LOWER(REPLACE(dt2.name,' ','')) 
              <= LOWER(REPLACE(dt.name,' ',''))
    ) AS doctype_order,

    s.name  AS serie_name,
    ss.name AS subserie_name,
    dt.name AS doctype_name

    FROM documental_studies ds

    JOIN dependencies d
        ON d.id = ds.dependency_id

    JOIN dependencies_function df
        ON df.id = ds.funciones_especificas

    JOIN series s 
        ON s.id = ds.serie_id

    JOIN subseries ss 
        ON ss.id = ds.subserie_id
    JOIN doctype dt 
        ON dt.id = ds.tipo_documental_id

    WHERE ds.entity_id = :entity_id

    ORDER BY
    series_order,
    subseries_order,
    doctype_order,
    df.function_number
    """


# =============================================================
#  NEW CODE GENERATOR 
# =============================================================

from collections import OrderedDict
from flask import request

@app.route("/tvd")
def tvd():

    entity_id = session.get("active_entity_id")

    entity = Entity.query.get(entity_id)

    studies = DocumentalStudy.query.filter_by(entity_id=entity_id).all()

    # reuse your estructura builder
    estructura = build_tvd_structure(studies, entity_id)

    dependency = None
    dependency_name = None

    if studies:
        dependency_name = studies[0].dependency.name if studies[0].dependency else None
    return render_template(
        "tvd_final.html",
        entity=entity,
        estructura=estructura,
        dependency=dependency
    )

# =========================================================
# APP ENTRYPOINT
# =========================================================
if __name__ == "__main__":
    try:
        with app.app_context():
            print("✅ Application initialized successfully.")
    except Exception as e:
        print("⚠️ Startup DB initialization failed:", e)
        print("➡️ Continuing with server startup anyway...")

    try:
        app.run(debug=True)
    except Exception as e:
        print("❌ Flask failed to start:", e)


